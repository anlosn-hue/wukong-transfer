# scripts/diting_excel.py
"""谛听采集表脚本。两条独立流程：
generate <批次JSON> <输出目录> —— 事后采集：生成按处室拆分的Excel（只含实际效果）+ 回写批次清单.json
parse <已填xlsx> <批次JSON> <输出JSON> —— 事后采集：解析已填表
generate-correction <批次JSON> <输出目录> —— 评估校正：生成板块式校正表（含风险等级校正行+F列行标识）+ 回写批次清单.json
parse-correction <已填xlsx> <批次JSON> <输出JSON> —— 评估校正：解析已填表
布局与批次JSON schema 见 docs/superpowers/plans/2026-07-04-谛听.md。
"""
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Protection
from openpyxl.worksheet.datavalidation import DataValidation

ENUM_IMPL = ["已采纳", "部分采纳", "未采纳"]
ENUM_LEVEL = ["维持原级", "一级", "二级", "三级", "四级"]
PART_B_FIELDS = {
    "活动": ["触达量", "实际参与人数", "转化率", "奖励核销情况",
             "投诉量及主要投诉点", "是否发生负面舆情及处置",
             "执行中是否临时调整规则", "处室自评与改进"],
    "系统维护": ["是否如期完成切换", "期间客户咨询与投诉量",
               "是否引发舆情", "应急预案是否启用", "处室自评与改进"],
}
_BRACKETS = re.compile(r"[（(【][^）)】]*[）)】]")  # 去内容只针对中文括号/()/【】；[ ] 留给 _ILLEGAL 只删符号保留内容
_QUOTES = re.compile(r"[\"'“”‘’「」『』]")
_ILLEGAL = re.compile(r"[:\\/?*\[\]]")


def derive_sheet_name(name, used):
    """活动名 → Excel sheet 名：去括号内容/引号/非法字符，取前12字；冲突加序号。"""
    s = _BRACKETS.sub("", name)
    s = _QUOTES.sub("", s)
    s = _ILLEGAL.sub("", s)
    s = s.replace(" ", "").strip()[:12] or "活动"
    base, n = s, 2
    while s in used:
        s = f"{base}{n}"
        n += 1
    return s


_BOLD = Font(bold=True)


# ========== 事后采集（业务处室填，按处室拆分，一活动一sheet，只含实际效果） ==========

def generate(batch, out_dir):
    """按处室生成多份 xlsx + 回写批次清单.json（含派生的 sheet 名）。返回回写后的 batch。"""
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    for f in batch["files"]:
        used = set()
        for act in f["activities"]:
            act["sheet"] = derive_sheet_name(act["name"], used)
            used.add(act["sheet"])
        wb = Workbook()
        overview = wb.active
        overview.title = "总览"
        _fill_overview(overview, batch, f)
        for act in f["activities"]:
            _fill_activity_sheet(wb.create_sheet(act["sheet"]), act)
        wb.save(out_dir / f["file"])
    manifest = out_dir / "批次清单.json"
    manifest.write_text(json.dumps(batch, ensure_ascii=False, indent=2), encoding="utf-8")
    return batch


def _fill_overview(ws, batch, f):
    ws.protection.sheet = True
    for col, w in (("A", 46), ("B", 26), ("C", 10), ("D", 16)):
        ws.column_dimensions[col].width = w
    ws["A1"] = f'{batch["batch_id"]} 活动事后反馈采集表（{f["dept"]}）'
    ws["A1"].font = _BOLD
    ws["A3"] = "填写说明：请在各活动分页「实际效果」栏填写；其余为预填信息（已锁定），请勿修改。"
    ws["A5"] = "本表包含活动："
    for c, v in zip("ABCD", ("活动名称", "起止时间", "类型", "对应分页")):
        ws[f"{c}6"] = v
        ws[f"{c}6"].font = _BOLD
    for i, act in enumerate(f["activities"]):
        r = 7 + i
        ws.cell(row=r, column=1, value=act["name"])
        ws.cell(row=r, column=2, value=f'{act.get("start_date", "")} ~ {act.get("end_date", "")}')
        ws.cell(row=r, column=3, value=act["category"])
        ws.cell(row=r, column=4, value=act["sheet"])


def _fill_activity_sheet(ws, act):
    ws.protection.sheet = True  # 防误改即可，不设密码
    for col, w in (("A", 24), ("B", 60)):
        ws.column_dimensions[col].width = w
    ws["A1"], ws["B1"] = "活动名称", act["name"]
    ws["A2"], ws["B2"] = "起止时间", f'{act.get("start_date", "")} ~ {act.get("end_date", "")}'
    ws["A3"], ws["B3"] = "类型", act["category"]
    ws["A5"] = "实际效果"
    ws["A5"].font = _BOLD
    for j, field in enumerate(PART_B_FIELDS[act["category"]]):
        r = 6 + j
        ws.cell(row=r, column=1, value=field)
        ws.cell(row=r, column=2).protection = Protection(locked=False)


def parse(xlsx_path, batch):
    """解析已填 xlsx。回传文件名随意，按 sheet 集合重合度选中批次清单里对应的 file 条目。"""
    wb = load_workbook(xlsx_path, data_only=True)
    entry = _select_file_entry(batch, wb)
    result = {"batch_id": batch["batch_id"], "file": entry["file"], "activities": [], "warnings": []}
    for act in entry["activities"]:
        ws, matched_by = _locate_sheet(wb, act)
        if ws is None:
            result["warnings"].append(f'{act["activity_id"]} 对应分页「{act["sheet"]}」未找到，整活动跳过')
            result["activities"].append({"activity_id": act["activity_id"], "matched_by": "missing", "effects": {}})
            continue
        effects, warnings = _parse_activity(ws, act)
        result["warnings"].extend(warnings)
        result["activities"].append({"activity_id": act["activity_id"], "matched_by": matched_by,
                                     "effects": effects})
    return result


def _select_file_entry(batch, wb):
    """选 sheet 命中数最多的 file 条目（exact 或 B1 活动名匹配都算命中）。零命中=拿错文件，直接报错。"""
    best, best_score = None, -1
    for f in batch["files"]:
        score = sum(1 for act in f["activities"] if _locate_sheet(wb, act)[0] is not None)
        if score > best_score:
            best, best_score = f, score
    if best_score == 0:
        raise ValueError("已填文件与本批次清单的任何条目都对不上，请核对文件/批次是否拿错")
    return best


def _locate_sheet(wb, act):
    """先按批次清单里的 sheet 名精确找；找不到再扫各分页 B1 活动名兜底（处室改了 sheet 名）。"""
    if act["sheet"] in wb.sheetnames:
        return wb[act["sheet"]], "exact"
    for name in wb.sheetnames:
        ws = wb[name]
        if ws["B1"].value == act["name"]:
            return ws, "fuzzy"
    return None, "missing"


def _cell_str(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return "" if v is None else str(v).strip()


def _parse_activity(ws, act):
    warnings = []
    effects = {}
    for j, field in enumerate(PART_B_FIELDS[act["category"]]):
        r = 6 + j
        if _cell_str(ws, r, 1) != field:  # 标签错位（处室插行等）→ A列全扫兜底
            r = _scan_label_row(ws, field)
            if r is None:
                warnings.append(f'{act["activity_id"]} 字段「{field}」未在分页中找到')
                effects[field] = ""
                continue
        effects[field] = _cell_str(ws, r, 2)
    return effects, warnings


def _scan_label_row(ws, label):
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == label:
            return r
    return None


# ========== 评估校正（舆情管理员内部填，板块式表，一活动一板块+风险等级校正行） ==========

def generate_correction(batch, out_dir):
    """板块式布局：每活动一板块（活动头行 + 风险等级校正行 + 建议行）+ 回写批次清单.json。

    识别设计（建议条数因活动而异，行号不可靠）：F 列为锁定的「行标识」——
    板块头 / 等级校正 / 建议1、建议2…（活动内编号）；生成时把标识同步回写进
    batch["rows"][i]["tag"]（就地变异传入的 batch），解析与网页展示都按
    E列活动ID + F列行标识定位。"""
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "评估建议采纳情况"
    ws.protection.sheet = True
    for col, w in (("A", 30), ("B", 60), ("C", 14), ("D", 40), ("E", 30), ("F", 10)):
        ws.column_dimensions[col].width = w
    ws["A1"] = f'{batch["batch_id"]} 评估建议采纳情况校正表'
    ws["A1"].font = _BOLD
    ws["A3"] = ('填写说明：按活动板块填写——「风险等级校正」行在C列选最终执行等级（未调整选'
                '"维持原级"），D列填调整原因；建议行在C列选是否采纳，D列可选填具体调整内容；'
                '其余列（含F列行标识）已锁定请勿修改。')
    for c, v in zip("ABCDEF", ("活动/建议", "内容", "校正/采纳", "说明", "活动ID", "行标识")):
        ws[f"{c}5"] = v
        ws[f"{c}5"].font = _BOLD
    dv_impl = DataValidation(type="list", formula1='"已采纳,部分采纳,未采纳"', allow_blank=True)
    dv_level = DataValidation(type="list", formula1='"维持原级,一级,二级,三级,四级"', allow_blank=True)
    ws.add_data_validation(dv_impl)
    ws.add_data_validation(dv_level)
    r = 6
    for aid, name, level, rows in _group_rows(batch):
        head = ws.cell(row=r, column=1, value=f"▍{name}")
        head.font = _BOLD
        ws.cell(row=r, column=5, value=aid)
        ws.cell(row=r, column=6, value="板块头")
        r += 1
        ws.cell(row=r, column=1, value="风险等级校正")
        ws.cell(row=r, column=2, value=f"原评估等级：{level or '（卡片未记录）'}")
        c_cell = ws.cell(row=r, column=3)
        c_cell.protection = Protection(locked=False)
        dv_level.add(c_cell)
        ws.cell(row=r, column=4).protection = Protection(locked=False)
        ws.cell(row=r, column=5, value=aid)
        ws.cell(row=r, column=6, value="等级校正")
        r += 1
        for idx, row in enumerate(rows, 1):
            row["tag"] = f"建议{idx}"   # 回写批次清单：解析按 (活动ID, 行标识) 定位
            ws.cell(row=r, column=2, value=row["suggestion"])
            c_cell = ws.cell(row=r, column=3)
            c_cell.protection = Protection(locked=False)
            dv_impl.add(c_cell)
            ws.cell(row=r, column=4).protection = Protection(locked=False)
            ws.cell(row=r, column=5, value=aid)
            ws.cell(row=r, column=6, value=row["tag"])
            r += 1
        r += 1   # 板块间空行
    file_name = f'校正表-{batch["batch_id"]}.xlsx'
    wb.save(out_dir / file_name)
    manifest = out_dir / "批次清单.json"
    manifest.write_text(json.dumps(batch, ensure_ascii=False, indent=2), encoding="utf-8")
    return file_name


def _group_rows(batch):
    """rows 按 activity_id 分组（保持首现顺序）；等级取顶层 levels 映射，缺省空串。"""
    order, groups = [], {}
    for row in batch["rows"]:
        aid = row["activity_id"]
        if aid not in groups:
            order.append(aid)
            groups[aid] = {"name": row["name"], "rows": []}
        groups[aid]["rows"].append(row)
    levels = batch.get("levels", {})
    return [(aid, groups[aid]["name"], levels.get(aid, ""), groups[aid]["rows"]) for aid in order]


def parse_correction(xlsx_path, batch):
    """解析板块式校正表。定位优先级（建议条数因活动而异，行号不可靠）：

    ① E列活动ID + F列行标识（generate 时已回写批次清单 rows[].tag）——主路径
    ② (activity_id, suggestion) 内容匹配——降级路径，兼容旧扁平格式表/旧批次清单
    标识命中但表内建议文本与批次清单不一致 → warning（锁定列被解锁改动），以批次清单文本为准。
    等级行按 F列=="等级校正"（或旧判据 A列=="风险等级校正"）+ 活动ID 定位；
    旧格式表无等级行 → warning + 空值，属预期降级。"""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    by_tag, by_content, level_index = {}, {}, {}
    for r in range(6, ws.max_row + 1):
        activity_id = _cell_str(ws, r, 5)
        if not activity_id:
            continue
        tag = _cell_str(ws, r, 6)
        if tag == "等级校正" or _cell_str(ws, r, 1) == "风险等级校正":
            level_index[activity_id] = (_cell_str(ws, r, 3), _cell_str(ws, r, 4))
            continue
        if tag == "板块头":
            continue
        vals = (_cell_str(ws, r, 3), _cell_str(ws, r, 4), _cell_str(ws, r, 2))
        if tag:
            by_tag[(activity_id, tag)] = vals
        if vals[2]:
            by_content[(activity_id, vals[2])] = vals
    result = {"batch_id": batch["batch_id"], "rows": [], "levels": {}, "warnings": []}
    for row in batch["rows"]:
        aid = row["activity_id"]
        hit = by_tag.get((aid, row.get("tag", ""))) or by_content.get((aid, row["suggestion"]))
        if hit is None:
            result["warnings"].append(
                f'{aid} 建议「{row["suggestion"][:20]}...」未在表中找到对应行，可能被误删'
            )
            result["rows"].append({"activity_id": aid, "suggestion": row["suggestion"],
                                   "落实情况": "", "说明": "", "dirty": False})
            continue
        impl, note, sheet_text = hit
        if row.get("tag") and sheet_text and sheet_text != row["suggestion"]:
            result["warnings"].append(
                f'{aid} {row["tag"]} 表内建议文本与批次清单不一致（锁定列疑被改动），以批次清单为准'
            )
        dirty = bool(impl) and impl not in ENUM_IMPL
        if dirty:
            result["warnings"].append(
                f'{aid} 建议「{row["suggestion"][:20]}...」落实情况为非标准值：{impl}'
            )
        result["rows"].append({"activity_id": aid, "suggestion": row["suggestion"],
                               "落实情况": impl, "说明": note, "dirty": dirty})
    for aid in dict.fromkeys(row["activity_id"] for row in batch["rows"]):
        if aid not in level_index:
            result["warnings"].append(f"{aid} 风险等级校正行未找到（旧格式表或被误删），等级按未校正处理")
            result["levels"][aid] = {"final": "", "note": "", "dirty": False}
            continue
        final, note = level_index[aid]
        dirty = bool(final) and final not in ENUM_LEVEL
        if dirty:
            result["warnings"].append(f"{aid} 风险等级校正为非标准值：{final}")
        result["levels"][aid] = {"final": final, "note": note, "dirty": dirty}
    return result


def _load_batch(path):
    p = Path(path)
    if not p.exists():
        print(f"批次清单不存在：{p}")
        sys.exit(1)
    return json.loads(p.read_text(encoding="utf-8"))


def main(argv):
    if len(argv) >= 3 and argv[0] == "generate":
        batch = _load_batch(argv[1])
        generate(batch, argv[2])
        print(f"已生成 {len(batch['files'])} 份采集表 + 批次清单.json → {argv[2]}")
    elif len(argv) >= 4 and argv[0] == "parse":
        batch = _load_batch(argv[2])
        result = parse(argv[1], batch)
        Path(argv[3]).write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"解析完成 → {argv[3]}（warnings: {len(result['warnings'])}）")
    elif len(argv) >= 3 and argv[0] == "generate-correction":
        batch = _load_batch(argv[1])
        filename = generate_correction(batch, argv[2])
        print(f"已生成评估校正表 {filename} + 批次清单.json → {argv[2]}")
    elif len(argv) >= 4 and argv[0] == "parse-correction":
        batch = _load_batch(argv[2])
        result = parse_correction(argv[1], batch)
        Path(argv[3]).write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"解析完成 → {argv[3]}（warnings: {len(result['warnings'])}）")
    else:
        print(__doc__)
        sys.exit(2)


if __name__ == "__main__":
    main(sys.argv[1:])
