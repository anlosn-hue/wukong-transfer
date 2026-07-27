# scripts/test_diting_excel.py
"""谛听采集表脚本测试。运行：python -m pytest scripts/test_diting_excel.py -v"""
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from openpyxl import load_workbook

from diting_excel import (
    ENUM_IMPL,
    PART_B_FIELDS,
    derive_sheet_name,
    generate,
    generate_correction,
    parse,
    parse_correction,
)


class TestDeriveSheetName:
    def test_removes_brackets_and_quotes(self):
        assert derive_sheet_name("粽叶飘香，端午安康（2026年手机银行端午节活动）", set()) == "粽叶飘香，端午安康"
        assert derive_sheet_name('"18"理财日', set()) == "18理财日"

    def test_truncates_to_12_chars(self):
        assert derive_sheet_name("第三方数字认证服务系统电子签名系统切换应急演练暂停服务", set()) == "第三方数字认证服务系统电"

    def test_strips_illegal_excel_chars(self):
        assert derive_sheet_name("A/B:C?D*E[F]G", set()) == "ABCDEFG"

    def test_conflict_appends_number(self):
        used = {"18理财日"}
        assert derive_sheet_name('"18"理财日', used) == "18理财日2"

    def test_empty_after_cleanup_falls_back(self):
        assert derive_sheet_name("（）", set()) == "活动"


def make_batch():
    """测试夹具：2处室3活动（活动类+系统维护类），无 suggestions 字段（v4起 Part A 完全移出此流程）。"""
    return {
        "batch_id": "2026-07-批次1",
        "created": "2026-07-04",
        "files": [
            {"file": "采集表-2026-07-批次1-策略经营处.xlsx", "dept": "策略经营处", "sent_date": "",
             "activities": [
                 {"activity_id": "20260611-策略经营处-端午节",
                  "name": "粽叶飘香，端午安康（2026年手机银行端午节活动）", "sheet": "",
                  "category": "活动", "start_date": "2026-06-18", "end_date": "2026-06-30"},
                 {"activity_id": "20260610-策略经营处-端午理财推送",
                  "name": "端午假期前理财产品营销宣传", "sheet": "",
                  "category": "活动", "start_date": "2026-06-10", "end_date": "2026-06-22"}]},
            {"file": "采集表-2026-07-批次1-平台建设处.xlsx", "dept": "平台建设处", "sent_date": "",
             "activities": [
                 {"activity_id": "20260617-平台建设处-手机银行城市服务暂停服务",
                  "name": "手机银行城市服务暂停服务（异地灾备切换演练）", "sheet": "",
                  "category": "系统维护", "start_date": "2026-06-27", "end_date": "2026-07-04"}]},
        ],
    }


class TestGenerate:
    def test_one_xlsx_per_dept_plus_manifest(self, tmp_path):
        batch = generate(make_batch(), tmp_path)
        assert (tmp_path / "采集表-2026-07-批次1-策略经营处.xlsx").exists()
        assert (tmp_path / "采集表-2026-07-批次1-平台建设处.xlsx").exists()
        manifest = json.loads((tmp_path / "批次清单.json").read_text(encoding="utf-8"))
        assert manifest["files"][0]["activities"][0]["sheet"] == "粽叶飘香，端午安康"

    def test_file_only_contains_own_dept_activities(self, tmp_path):
        generate(make_batch(), tmp_path)
        wb = load_workbook(tmp_path / "采集表-2026-07-批次1-平台建设处.xlsx")
        assert wb.sheetnames == ["总览", "手机银行城市服务暂停服务"]

    def test_activity_sheet_layout(self, tmp_path):
        generate(make_batch(), tmp_path)
        wb = load_workbook(tmp_path / "采集表-2026-07-批次1-策略经营处.xlsx")
        ws = wb["粽叶飘香，端午安康"]
        assert ws["B1"].value == "粽叶飘香，端午安康（2026年手机银行端午节活动）"
        assert ws["A5"].value == "实际效果"
        assert ws["A6"].value == "触达量"
        assert ws["A13"].value == "处室自评与改进"

    def test_maintenance_category_uses_own_fields(self, tmp_path):
        generate(make_batch(), tmp_path)
        wb = load_workbook(tmp_path / "采集表-2026-07-批次1-平台建设处.xlsx")
        ws = wb["手机银行城市服务暂停服务"]
        assert ws["A6"].value == "是否如期完成切换"
        assert ws["A10"].value == "处室自评与改进"

    def test_protection_and_unlocked_cells(self, tmp_path):
        generate(make_batch(), tmp_path)
        wb = load_workbook(tmp_path / "采集表-2026-07-批次1-策略经营处.xlsx")
        ws = wb["粽叶飘香，端午安康"]
        assert ws.protection.sheet is True
        assert ws["A6"].protection.locked is True    # 字段名锁定
        assert ws["B6"].protection.locked is False   # 填写值可填

    def test_overview_sheet(self, tmp_path):
        generate(make_batch(), tmp_path)
        wb = load_workbook(tmp_path / "采集表-2026-07-批次1-策略经营处.xlsx")
        ws = wb["总览"]
        assert "策略经营处" in ws["A1"].value
        assert ws["A7"].value == "粽叶飘香，端午安康（2026年手机银行端午节活动）"


def _fill_in(tmp_path, filename, writes):
    """模拟处室填表：writes = [(sheet, cell, value), ...]"""
    p = tmp_path / filename
    wb = load_workbook(p)
    for sheet, cell, value in writes:
        wb[sheet][cell] = value
    wb.save(p)
    return p


class TestParse:
    def test_round_trip(self, tmp_path):
        batch = generate(make_batch(), tmp_path)
        p = _fill_in(tmp_path, "采集表-2026-07-批次1-策略经营处.xlsx", [
            ("粽叶飘香，端午安康", "B6", "推送量120万"),
            ("粽叶飘香，端午安康", "B13", "整体平稳，下次提前备货"),
            ("端午假期前理财产品营销宣", "B6", "触达8.6万人"),
        ])
        result = parse(p, batch)
        acts = {a["activity_id"]: a for a in result["activities"]}
        duanwu = acts["20260611-策略经营处-端午节"]
        assert duanwu["matched_by"] == "exact"
        assert duanwu["effects"]["触达量"] == "推送量120万"
        assert duanwu["effects"]["处室自评与改进"] == "整体平稳，下次提前备货"
        assert duanwu["effects"]["转化率"] == ""
        push = acts["20260610-策略经营处-端午理财推送"]
        assert push["effects"]["触达量"] == "触达8.6万人"
        assert result["warnings"] == []

    def test_selects_correct_file_entry(self, tmp_path):
        batch = generate(make_batch(), tmp_path)
        src = tmp_path / "采集表-2026-07-批次1-平台建设处.xlsx"
        renamed = tmp_path / "回传-随便起的名字.xlsx"
        renamed.write_bytes(src.read_bytes())
        result = parse(renamed, batch)
        ids = [a["activity_id"] for a in result["activities"]]
        assert ids == ["20260617-平台建设处-手机银行城市服务暂停服务"]

    def test_maintenance_fields(self, tmp_path):
        batch = generate(make_batch(), tmp_path)
        p = _fill_in(tmp_path, "采集表-2026-07-批次1-平台建设处.xlsx", [
            ("手机银行城市服务暂停服务", "B6", "是，7月4日18:00如期恢复"),
        ])
        result = parse(p, batch)
        act = result["activities"][0]
        assert act["effects"]["是否如期完成切换"] == "是，7月4日18:00如期恢复"
        assert set(act["effects"].keys()) == set(PART_B_FIELDS["系统维护"])


class TestParseTolerance:
    def test_renamed_sheet_fuzzy_matched(self, tmp_path):
        batch = generate(make_batch(), tmp_path)
        p = tmp_path / "采集表-2026-07-批次1-平台建设处.xlsx"
        wb = load_workbook(p)
        wb["手机银行城市服务暂停服务"].title = "城市服务反馈"
        wb.save(p)
        result = parse(p, batch)
        act = result["activities"][0]
        assert act["matched_by"] == "fuzzy"
        assert set(act["effects"].keys()) == set(PART_B_FIELDS["系统维护"])

    def test_deleted_sheet_reported_missing(self, tmp_path):
        batch = generate(make_batch(), tmp_path)
        p = tmp_path / "采集表-2026-07-批次1-策略经营处.xlsx"
        wb = load_workbook(p)
        del wb["端午假期前理财产品营销宣"]
        wb.save(p)
        result = parse(p, batch)
        acts = {a["activity_id"]: a for a in result["activities"]}
        assert acts["20260610-策略经营处-端午理财推送"]["matched_by"] == "missing"
        assert any("未找到" in w for w in result["warnings"])
        assert acts["20260611-策略经营处-端午节"]["matched_by"] == "exact"

    def test_wrong_workbook_rejected(self, tmp_path):
        import pytest
        from openpyxl import Workbook
        batch = generate(make_batch(), tmp_path)
        stray = tmp_path / "无关文件.xlsx"
        wb = Workbook()
        wb.active.title = "随便什么表"
        wb.save(stray)
        with pytest.raises(ValueError):
            parse(stray, batch)


def make_correction_batch():
    """评估校正测试夹具：2活动3条建议，扁平结构。"""
    return {
        "batch_id": "2026-07-批次1",
        "created": "2026-07-04",
        "rows": [
            {"activity_id": "20260611-策略经营处-端午节", "name": "端午节",
             "suggestion": "免责声明弱化第三方推卸表述", "round": 1},
            {"activity_id": "20260611-策略经营处-端午节", "name": "端午节",
             "suggestion": "标注中奖概率", "round": 1},
            {"activity_id": "20260610-策略经营处-端午理财推送", "name": "端午理财推送",
             "suggestion": "补交消保审核凭证", "round": 1},
        ],
    }


class TestGenerateCorrection:
    """板块式布局（2026-07 改造后）：每活动一板块 = 板块头 + 风险等级校正行 + 建议行 N 条，
    板块间空一行。F 列「行标识」是解析定位的锚——建议条数因活动而异，行号不可靠。

    标准夹具（2活动/3建议）的实际坐标：
        5 表头 | 6 板块头(端午节) | 7 等级校正 | 8 建议1 | 9 建议2 | 10 空
        11 板块头(端午理财推送) | 12 等级校正 | 13 建议1
    """

    def test_block_layout(self, tmp_path):
        batch = make_correction_batch()
        filename = generate_correction(batch, tmp_path)
        assert (tmp_path / filename).exists()
        assert (tmp_path / "批次清单.json").exists()
        ws = load_workbook(tmp_path / filename).active
        assert ws.title == "评估建议采纳情况"
        assert [ws[f"{c}5"].value for c in "ABCDEF"] ==             ["活动/建议", "内容", "校正/采纳", "说明", "活动ID", "行标识"]
        # 板块一
        assert ws["A6"].value == "▍端午节"
        assert ws["F6"].value == "板块头"
        assert ws["A7"].value == "风险等级校正" and ws["F7"].value == "等级校正"
        assert ws["B8"].value == "免责声明弱化第三方推卸表述" and ws["F8"].value == "建议1"
        assert ws["B9"].value == "标注中奖概率" and ws["F9"].value == "建议2"
        assert ws["A8"].value is None, "建议行 A 列留空，活动名只出现在板块头"
        # 板块间空一行，板块二从 11 起
        assert ws["A10"].value is None
        assert ws["A11"].value == "▍端午理财推送"
        assert ws["B13"].value == "补交消保审核凭证" and ws["F13"].value == "建议1"
        # E 列活动ID 每行都带，供解析定位
        assert ws["E8"].value == "20260611-策略经营处-端午节"
        assert ws["E13"].value == "20260610-策略经营处-端午理财推送"

    def test_level_row_shows_original_level(self, tmp_path):
        batch = make_correction_batch()
        batch["levels"] = {"20260611-策略经营处-端午节": "三级"}
        ws = load_workbook(tmp_path / generate_correction(batch, tmp_path)).active
        assert ws["B7"].value == "原评估等级：三级"
        assert ws["B12"].value == "原评估等级：（卡片未记录）", "levels 缺项要显式说明，不留空"

    def test_tags_written_back_to_batch(self, tmp_path):
        """生成时把行标识就地回写进 batch，解析按 (活动ID, 行标识) 定位靠它。"""
        batch = make_correction_batch()
        generate_correction(batch, tmp_path)
        assert [r["tag"] for r in batch["rows"]] == ["建议1", "建议2", "建议1"]

    def test_protection(self, tmp_path):
        batch = make_correction_batch()
        ws = load_workbook(tmp_path / generate_correction(batch, tmp_path)).active
        assert ws.protection.sheet is True
        # 板块头整行锁定——那行没有可填内容
        assert ws["C6"].protection.locked is True
        assert ws["D6"].protection.locked is True
        # 等级校正行与建议行：只解锁 C（选值）与 D（说明）
        for row in (7, 8, 9, 12, 13):
            assert ws[f"C{row}"].protection.locked is False, f"C{row} 应可填"
            assert ws[f"D{row}"].protection.locked is False, f"D{row} 应可填"
            for col in "ABEF":
                assert ws[f"{col}{row}"].protection.locked is True, f"{col}{row} 应锁定"


def _fill_correction(tmp_path, filename, writes):
    """模拟管理员填评估校正表：writes = [(cell, value), ...]"""
    p = tmp_path / filename
    wb = load_workbook(p)
    ws = wb.active
    for cell, value in writes:
        ws[cell] = value
    wb.save(p)
    return p


class TestParseCorrection:
    def test_round_trip(self, tmp_path):
        batch = make_correction_batch()
        filename = generate_correction(batch, tmp_path)
        p = _fill_correction(tmp_path, filename, [
            ("C8", "已采纳"), ("D8", "已按建议修改"),
            ("C9", "未采纳"),
            ("C13", "部分采纳"), ("D13", "补了口头说明，未存档"),
        ])
        result = parse_correction(p, batch)
        rows = result["rows"]
        assert rows[0]["落实情况"] == "已采纳"
        assert rows[0]["说明"] == "已按建议修改"
        assert rows[0]["dirty"] is False
        assert rows[1]["落实情况"] == "未采纳"
        assert rows[2]["activity_id"] == "20260610-策略经营处-端午理财推送"
        assert rows[2]["落实情况"] == "部分采纳"
        assert result["warnings"] == []

    def test_level_correction_parsed(self, tmp_path):
        """等级校正是板块式新增的一整块能力，此前无任何测试覆盖。"""
        batch = make_correction_batch()
        filename = generate_correction(batch, tmp_path)
        p = _fill_correction(tmp_path, filename, [
            ("C7", "四级"), ("D7", "破净背景缓解，回落"),
            ("C12", "维持原级"),
        ])
        levels = parse_correction(p, batch)["levels"]
        assert levels["20260611-策略经营处-端午节"]["final"] == "四级"
        assert levels["20260611-策略经营处-端午节"]["note"] == "破净背景缓解，回落"
        assert levels["20260611-策略经营处-端午节"]["dirty"] is False
        assert levels["20260610-策略经营处-端午理财推送"]["final"] == "维持原级"

    def test_level_dirty_value_flagged(self, tmp_path):
        """把建议行的取值填进等级校正行——取值域不同，必须逮住。"""
        batch = make_correction_batch()
        filename = generate_correction(batch, tmp_path)
        p = _fill_correction(tmp_path, filename, [("C7", "已采纳")])
        result = parse_correction(p, batch)
        lv = result["levels"]["20260611-策略经营处-端午节"]
        assert lv["dirty"] is True
        assert any("风险等级校正为非标准值" in w for w in result["warnings"])

    def test_dirty_value_flagged(self, tmp_path):
        batch = make_correction_batch()
        filename = generate_correction(batch, tmp_path)
        p = _fill_correction(tmp_path, filename, [("C8", "基本采纳")])
        result = parse_correction(p, batch)
        assert result["rows"][0]["落实情况"] == "基本采纳"
        assert result["rows"][0]["dirty"] is True
        assert any("非标准值" in w for w in result["warnings"])

    def test_empty_rows_batch_not_rendered(self, tmp_path):
        """建议为空的批次（rows=[]）不生成任何数据行、不报错。"""
        batch = {"batch_id": "2026-07-批次2", "created": "2026-07-04", "rows": []}
        filename = generate_correction(batch, tmp_path)
        ws = load_workbook(tmp_path / filename).active
        assert ws["A6"].value is None

    def test_deleted_row_reported_missing(self, tmp_path):
        """管理员误删一行 → 该条 warning，其余行不受影响（按 活动ID+行标识 定位，不依赖行位置）。"""
        batch = make_correction_batch()
        filename = generate_correction(batch, tmp_path)
        p = tmp_path / filename
        wb = load_workbook(p)
        ws = wb.active
        ws.delete_rows(9, 1)  # 删掉"标注中奖概率"（建议2）
        wb.save(p)
        result = parse_correction(p, batch)
        assert any("未在表中找到" in w for w in result["warnings"])
        pushed = [r for r in result["rows"]
                  if r["activity_id"] == "20260610-策略经营处-端午理财推送"][0]
        assert pushed["dirty"] is False
        assert pushed["落实情况"] == ""

    def test_same_name_different_id_not_confused(self, tmp_path):
        """同名不同ID的两个活动，验证 parse_correction 不靠活动名做任何匹配/兜底。

        该夹具每活动仅1条建议，坐标为：6 板块头甲 / 7 等级校正 / 8 建议A /
        9 空 / 10 板块头乙 / 11 等级校正 / 12 建议B。
        """
        batch = {
            "batch_id": "2026-07-批次3", "created": "2026-07-04",
            "rows": [
                {"activity_id": "20260101-甲处-同名活动", "name": "同名活动",
                 "suggestion": "建议A", "round": 1},
                {"activity_id": "20260201-乙处-同名活动", "name": "同名活动",
                 "suggestion": "建议B", "round": 1},
            ],
        }
        filename = generate_correction(batch, tmp_path)
        p = _fill_correction(tmp_path, filename, [("C8", "已采纳"), ("C12", "未采纳")])
        result = parse_correction(p, batch)
        rows = {r["suggestion"]: r for r in result["rows"]}
        assert rows["建议A"]["落实情况"] == "已采纳"
        assert rows["建议B"]["落实情况"] == "未采纳"
