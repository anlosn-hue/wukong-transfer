#!/usr/bin/env python3
"""
声誉风险事前评估表生成器
用法：python gen_eval_table.py --template <xlsx路径> --output <xlsx路径> --data <json路径>

data JSON 字段：
    title         str  评估表标题（如 "XX活动声誉风险事前评估表"）
    dept          str  部门/机构名称
    service_name  str  业务/产品或服务名称
    section1      str  一、评估事项 内容
    feasibility   str  二、声誉风险评估 — 可行性分析
    public_acceptance str  民众认可度分析
    public_opinion    str  舆论关注度分析
    risk_prob     str  风险发生概率（"低"/"中"/"高"）
    risk_level    str  声誉风险等级（活动/系统类均填占位符）
    result_usage  str  评估结果运用
    section3      str  三、产品宣传环节及处置预案 内容
    section4      str  四、媒体备答口径 内容
"""
import json
import argparse
from pathlib import Path
import openpyxl


def _find_row(ws, col_idx: int, pattern: str):
    """在指定列中查找包含 pattern 的第一行，返回行号（1-indexed）或 None。"""
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=col_idx).value
        if v and pattern in str(v):
            return r
    return None


def _safe_write(ws, row: int, col: int, value):
    """写入单元格，自动处理合并单元格（写入合并区域的锚点）。"""
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            ws.cell(row=mr.min_row, column=mr.min_col).value = value
            return
    ws.cell(row=row, column=col).value = value


def fill_eval_table(template_path: str, output_path: str, data: dict) -> None:
    wb = openpyxl.load_workbook(template_path)
    if '声誉风险事前评估表' not in wb.sheetnames:
        raise ValueError(f'模板缺少工作表「声誉风险事前评估表」，实际工作表：{wb.sheetnames}')
    ws = wb['声誉风险事前评估表']

    FILL_RULES = [
        # (搜索列, 搜索关键词, 填入列, data_key)
        (1, '声誉风险事前评估表', 1, 'title'),
        (1, '部门/机构名称',       2, 'dept'),
        (1, '评估内容',            2, 'service_name'),
        (1, '一、评估事项',        2, 'section1'),
        (4, '阐述可行性',          3, 'feasibility'),
        (4, '阐述群众认可度',      3, 'public_acceptance'),
        (4, '阐述舆论关注度',      3, 'public_opinion'),
        (4, '评估此项业务',        3, 'risk_prob'),
        (2, '根据本行声誉风险等级分类', 3, 'risk_level'),
        (2, '评估结果运用',        3, 'result_usage'),
        (1, '三、产品宣传',        2, 'section3'),
        (1, '四、媒体备答口径',    2, 'section4'),
    ]

    for search_col, pattern, fill_col, key in FILL_RULES:
        r = _find_row(ws, search_col, pattern)
        if r is not None and key in data:
            _safe_write(ws, r, fill_col, data[key])
        elif r is None:
            print(f'[WARN] 未找到锚点「{pattern}」（搜索列 {search_col}），跳过字段 {key}')

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f'[OK] 已生成评估表：{output_path}')


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='生成声誉风险事前评估表')
    parser.add_argument('--template', required=True, help='模板 xlsx 路径')
    parser.add_argument('--output',   required=True, help='输出 xlsx 路径')
    parser.add_argument('--data',     required=True, help='数据 JSON 文件路径')
    args = parser.parse_args()

    with open(args.data, encoding='utf-8-sig') as f:
        data = json.load(f)

    fill_eval_table(args.template, args.output, data)
