# -*- coding: utf-8 -*-
"""label2xlsx.py 闸门与合成测试。"""
import json
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

SCRIPT = Path(__file__).resolve().parents[1] / 'label2xlsx.py'
sys.path.insert(0, str(SCRIPT.parent))
import label2xlsx  # noqa: E402

YES = {"row": 2, "problem": "是", "labels": ["保本保息或收益承诺"],
       "originals": ["本金相对安全的"], "reasons": ["经理对本金安全作肯定表述"],
       "confidence": ["2级"], "suppressed": None}
NO = {"row": 3, "problem": "否", "labels": [], "originals": [],
      "reasons": ["仅说明合同约定利率"], "confidence": [], "suppressed": None}


def prep(tmp_path, ann, mapping=None):
    if mapping is None:
        mapping = [{"row": r["row"], "raw": "原文%d" % r["row"]} for r in ann]
    for name, recs in (('04-标注.jsonl', ann), ('01-原文映射.jsonl', mapping)):
        (tmp_path / name).write_text(
            '\n'.join(json.dumps(r, ensure_ascii=False) for r in recs) + '\n',
            encoding='utf-8')
    return tmp_path


def run(batch):
    return subprocess.run([sys.executable, str(SCRIPT), str(batch)],
                          capture_output=True)


def test_happy_path_two_rows(tmp_path):
    prep(tmp_path, [NO, YES])           # 故意乱序，验证按行号排序
    p = run(tmp_path)
    assert p.returncode == 0
    ws = load_workbook(tmp_path / '04-合规预警标注表.xlsx')['标注表']
    assert [c.value for c in ws[1]] == ['聊天内容原文', '是否存在问题', '违规预警标签',
                                        '问题原文', '判断理由', '风险判断可信度等级']
    assert ws.cell(2, 1).value == '原文2' and ws.cell(2, 2).value == '是'
    assert ws.cell(2, 3).value == '保本保息或收益承诺'
    assert ws.cell(2, 4).value == '1.保本保息或收益承诺：本金相对安全的'
    assert ws.cell(2, 6).value == '2级'
    assert ws.cell(3, 2).value == '否' and ws.cell(3, 3).value is None
    assert ws.cell(3, 5).value == '仅说明合同约定利率'
    assert ws.freeze_panes == 'A2' and ws.auto_filter.ref


def test_multi_label_join(tmp_path):
    y = dict(YES, labels=["风险提示不到位或弱化风险", "保本保息或收益承诺"],
             originals=["比较安全", "本金相对安全的"], reasons=["理1", "理2"],
             confidence=["2级", "2级"])
    prep(tmp_path, [y])
    assert run(tmp_path).returncode == 0
    ws = load_workbook(tmp_path / '04-合规预警标注表.xlsx')['标注表']
    assert ws.cell(2, 3).value == '风险提示不到位或弱化风险；保本保息或收益承诺'
    assert ws.cell(2, 4).value == ('1.风险提示不到位或弱化风险：比较安全；'
                                   '2.保本保息或收益承诺：本金相对安全的')
    assert ws.cell(2, 6).value == '2级；2级'


@pytest.mark.parametrize('mutate,msg', [
    (lambda a: a.append(dict(YES)), '行号重复'),
    (lambda a: a.pop(1), '缺行'),
    (lambda a: a.append(dict(YES, row=99)), '多行'),
    (lambda a: a.__setitem__(0, dict(YES, problem='维持')), 'problem'),
    (lambda a: a.__setitem__(0, dict(YES, labels=['骚扰营销或无视客户拒绝'])), '非法标签'),
    (lambda a: a.__setitem__(0, dict(YES, confidence=['六级'])), '可信度'),
    (lambda a: a.__setitem__(0, dict(YES, reasons=['理1', '理2'])), '等长'),
    (lambda a: a.__setitem__(0, dict(YES, originals=[''])), '空字符串'),
    (lambda a: a.__setitem__(1, dict(NO, labels=['饥饿营销'])), '判否行'),
])
def test_gates_block_and_no_xlsx(tmp_path, mutate, msg):
    ann = [dict(YES), dict(NO)]
    mapping = [{"row": 2, "raw": "原文2"}, {"row": 3, "raw": "原文3"}]
    mutate(ann)
    prep(tmp_path, ann, mapping)
    p = run(tmp_path)
    assert p.returncode == 1
    assert msg in p.stdout.decode('utf-8')
    assert not (tmp_path / '04-合规预警标注表.xlsx').exists()


def test_array_field_type_gate(tmp_path):
    prep(tmp_path, [dict(YES, labels='保本保息或收益承诺'), dict(NO)],
         [{"row": 2, "raw": "原文2"}, {"row": 3, "raw": "原文3"}])
    p = run(tmp_path)
    assert p.returncode == 1 and '须为数组' in p.stdout.decode('utf-8')


def test_mapping_duplicate_row_gate(tmp_path):
    prep(tmp_path, [dict(YES)],
         [{"row": 2, "raw": "a"}, {"row": 2, "raw": "b"}])
    p = run(tmp_path)
    assert p.returncode == 1 and '原文映射行号重复' in p.stdout.decode('utf-8')


def test_missing_ann_file_gate(tmp_path):
    # 只写原文映射，不写 04-标注.jsonl -> 裸 traceback 应被 say() 通道拦截
    (tmp_path / '01-原文映射.jsonl').write_text(
        json.dumps({"row": 2, "raw": "原文2"}, ensure_ascii=False) + '\n',
        encoding='utf-8')
    p = run(tmp_path)
    assert p.returncode == 1
    assert '格式异常' in p.stdout.decode('utf-8')
    assert not (tmp_path / '04-合规预警标注表.xlsx').exists()


def test_fit_height_caps_at_380_for_long_text():
    texts = ['字' * 2000, None, None, None, None, None]
    widths = [80, 10, 28, 45, 55, 16]
    assert label2xlsx.fit_height(texts, widths) == 380


def test_fit_height_uncapped_formula_for_short_text():
    texts = ['短文本', None, None, None, None, None]
    widths = [80, 10, 28, 45, 55, 16]
    size = 10.5
    n = max(1, -(-label2xlsx.vlen('短文本') // (widths[0] - 2)))
    expected = n * (size + 3.5) + 4
    assert expected < 380
    assert label2xlsx.fit_height(texts, widths) == expected
