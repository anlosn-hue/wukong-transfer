# -*- coding: utf-8 -*-
"""md2xlsx.py convert_03 的六/五 sheet 测试。fixture 取真实批次 03 md 改造。"""
import re
import shutil
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

SCRIPT = Path(__file__).resolve().parents[1] / 'md2xlsx.py'
REAL = Path(r'work/事后质检/报告'
            r'\20260726-测试样例')

pytestmark = pytest.mark.skipif(not REAL.exists(),
                                reason='数据区真实批次不在本机（junction 未接）')

SEC7 = '''
## 七、标注结论与质检结论差异核对

| 行号 | 标注结论 | 质检结论 | 差异归因 | 建议复核方向 |
|---|---|---|---|---|
| 12 | 否 | 维度①命中（中） | 证据不足 | 复核该会话上下文是否完整、标注判否是否成立 |
'''

SEC7_ZERO_NOTE = '''
## 七、标注结论与质检结论差异核对

本批标注结论与质检结论一致，未发现实质差异。
'''

SEC7_NO_NOTE = '''
## 七、标注结论与质检结论差异核对
'''


def prep(tmp_path, with_sec7, sec7_text=None):
    for f in ('02-审计关注点清单.md', '03-体检报告.md'):
        shutil.copy(REAL / f, tmp_path / f)
    if with_sec7:
        text = sec7_text if sec7_text is not None else SEC7
        md = tmp_path / '03-体检报告.md'
        lines = md.read_text(encoding='utf-8').splitlines()
        # 第七节插在落款（末两行非空行）之前
        idx = max(i for i, l in enumerate(lines) if l.strip()) - 1
        lines[idx:idx] = text.splitlines()
        md.write_text('\n'.join(lines), encoding='utf-8')
    return tmp_path


def run(batch):
    return subprocess.run([sys.executable, str(SCRIPT), str(batch)],
                          capture_output=True)


def strip_voice_table(tmp_path):
    """把第四节（语音盲区）表格整段替换为一行「无待抽听项」，供区间越界回归测试复用"""
    md = tmp_path / '03-体检报告.md'
    lines = md.read_text(encoding='utf-8').splitlines()
    sec4 = lines.index('## 四、语音盲区清单（待人工抽听）')
    sec5 = lines.index('## 五、待人工确认项')
    lines[sec4 + 1:sec5] = ['', '无待抽听项', '']
    md.write_text('\n'.join(lines), encoding='utf-8')


SPOT_CHECK_SUBSECTION = '''
### 按标注纪律判否的抽检建议
1. 行108：疑似弱线索上下文缺失
- 行109：另一条抽检线索
'''


def insert_spot_check_subsection(tmp_path):
    """在第五节编号项之后、## 六、沉淀建议 之前插入抽检建议小节（含误用编号格式的诱饵行）"""
    md = tmp_path / '03-体检报告.md'
    lines = md.read_text(encoding='utf-8').splitlines()
    sec6 = lines.index('## 六、沉淀建议')
    lines[sec6:sec6] = SPOT_CHECK_SUBSECTION.splitlines()
    md.write_text('\n'.join(lines), encoding='utf-8')


def test_with_sec7_six_sheets(tmp_path):
    prep(tmp_path, with_sec7=True)
    assert run(tmp_path).returncode == 0
    wb = load_workbook(tmp_path / '03-体检报告.xlsx')
    assert wb.sheetnames == ['总览', '七维明细', '语音盲区待抽听',
                             '待人工确认', '沉淀建议', '差异核对']
    ws = wb['差异核对']
    assert ws.cell(1, 1).value == '行号' and ws.cell(2, 4).value == '证据不足'


def test_without_sec7_backward_compatible(tmp_path):
    prep(tmp_path, with_sec7=False)
    assert run(tmp_path).returncode == 0
    wb = load_workbook(tmp_path / '03-体检报告.xlsx')
    assert '差异核对' not in wb.sheetnames and len(wb.sheetnames) == 5


def test_zero_diff_with_note(tmp_path):
    prep(tmp_path, with_sec7=True, sec7_text=SEC7_ZERO_NOTE)
    assert run(tmp_path).returncode == 0
    wb = load_workbook(tmp_path / '03-体检报告.xlsx')
    ws = wb['差异核对']
    assert ws.cell(1, 1).value == '本批标注结论与质检结论一致，未发现实质差异。'


def test_zero_diff_without_note_falls_back(tmp_path):
    prep(tmp_path, with_sec7=True, sec7_text=SEC7_NO_NOTE)
    assert run(tmp_path).returncode == 0
    wb = load_workbook(tmp_path / '03-体检报告.xlsx')
    ws = wb['差异核对']
    assert ws.cell(1, 1).value == '本批无实质差异'
    assert ws.cell(1, 1).value != '数字运营部声誉风险管理智能体·悟空'


def test_voice_section_without_table_does_not_steal_sec7_table(tmp_path):
    prep(tmp_path, with_sec7=True)          # 复用现有 fixture 的带表 SEC7
    strip_voice_table(tmp_path)             # 第四节清空为无表状态
    result = run(tmp_path)
    assert result.returncode == 0, result.stderr.decode('utf-8', 'ignore')
    wb = load_workbook(tmp_path / '03-体检报告.xlsx')

    ws3 = wb['语音盲区待抽听']
    assert ws3.cell(1, 1).value != '行号'
    voice_text = ' '.join(str(c.value) for row in ws3.iter_rows()
                          for c in row if c.value is not None)
    assert '证据不足' not in voice_text

    ws6 = wb['差异核对']
    assert ws6.cell(2, 4).value == '证据不足'


def test_spot_check_subsection_not_swallowed(tmp_path):
    prep(tmp_path, with_sec7=False)
    md = tmp_path / '03-体检报告.md'
    lines = md.read_text(encoding='utf-8').splitlines()
    sec5 = lines.index('## 五、待人工确认项')
    sec6 = lines.index('## 六、沉淀建议')
    expected_items = sum(1 for l in lines[sec5 + 1:sec6]
                         if re.match(r'\d+\.\s+', l.strip()))

    insert_spot_check_subsection(tmp_path)
    result = run(tmp_path)
    assert result.returncode == 0, result.stderr.decode('utf-8', 'ignore')
    wb = load_workbook(tmp_path / '03-体检报告.xlsx')
    ws4 = wb['待人工确认']
    assert ws4.max_row == expected_items + 1     # +1 表头行，不含 108/109 任何一条
    text = ' '.join(str(c.value) for row in ws4.iter_rows()
                    for c in row if c.value is not None)
    assert '行108' not in text and '行109' not in text
