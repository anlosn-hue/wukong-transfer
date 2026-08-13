# -*- coding: utf-8 -*-
"""parse_chats.py 的 sheet 选择/表头/原文映射测试。fixture 用 openpyxl 现造。"""
import json
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook

SCRIPT = Path(__file__).resolve().parents[1] / 'parse_chats.py'
CHAT = '张经理102938 10:00:01：  这款理财收益稳。@@@王姐 10:00:30：  好的'


def make_xlsx(path, sheets):
    """sheets: [(名, A1, [会话行...])]"""
    wb = Workbook()
    wb.remove(wb.active)
    for name, a1, rows in sheets:
        ws = wb.create_sheet(name)
        if a1 is not None:
            ws.cell(1, 1, a1)
        for i, r in enumerate(rows, 2 if a1 is not None else 1):
            ws.cell(i, 1, r)
    wb.save(path)


def run(xlsx, outdir):
    return subprocess.run([sys.executable, str(SCRIPT), str(xlsx), str(outdir)],
                          capture_output=True)


def stats_of(proc):
    return json.loads(proc.stdout.decode('utf-8'))


def test_single_match_picks_that_sheet(tmp_path):
    x = tmp_path / 'a.xlsx'
    make_xlsx(x, [('说明页', '规则说明', []),
                  ('数据', '聊天内容', [CHAT])])
    p = run(x, tmp_path / 'out')
    assert p.returncode == 0
    s = stats_of(p)
    assert s['所用sheet'] == '数据' and s['会话数'] == 1
    assert '说明页' in s['其余sheet']


def test_content_header_matches_case_insensitive(tmp_path):
    x = tmp_path / 'b.xlsx'
    make_xlsx(x, [('Sheet1 (2)', ' Content ', [CHAT])])
    p = run(x, tmp_path / 'out')
    assert p.returncode == 0
    assert stats_of(p)['会话数'] == 1  # 表头被识别跳过，不算会话


def test_zero_match_single_sheet_used_with_flag(tmp_path):
    x = tmp_path / 'c.xlsx'
    make_xlsx(x, [('Sheet1', CHAT, [CHAT])])  # A1 直接是会话
    p = run(x, tmp_path / 'out')
    assert p.returncode == 0
    s = stats_of(p)
    assert s['表头异常'] is True and s['会话数'] == 2


def test_zero_match_multi_sheet_errors(tmp_path):
    x = tmp_path / 'd.xlsx'
    make_xlsx(x, [('甲', '别的', [CHAT]), ('乙', '别的2', [CHAT])])
    p = run(x, tmp_path / 'out')
    assert p.returncode == 1
    assert '请用户指定' in p.stdout.decode('utf-8')


def test_multi_match_errors(tmp_path):
    x = tmp_path / 'e.xlsx'
    make_xlsx(x, [('甲', '聊天内容', [CHAT]), ('乙', 'content', [CHAT])])
    p = run(x, tmp_path / 'out')
    assert p.returncode == 1
    assert '请用户指定' in p.stdout.decode('utf-8')


def test_raw_mapping_jsonl(tmp_path):
    x = tmp_path / 'f.xlsx'
    make_xlsx(x, [('数据', '聊天内容', [CHAT, None, CHAT])])  # 中间空行
    out = tmp_path / 'out'
    p = run(x, out)
    assert p.returncode == 0
    recs = [json.loads(l) for l in
            (out / '01-原文映射.jsonl').read_text(encoding='utf-8').splitlines()]
    assert [r['row'] for r in recs] == [2, 4]      # 空行 3 不进映射
    assert recs[0]['raw'] == CHAT                   # 原文逐字保留


def test_sheet_parse_failure_friendly_error(tmp_path, monkeypatch, capfdbinary):
    """选中 sheet 后全量读取失败 -> 友好报错 exit 1（b455fe7 防御分支）。"""
    import pandas as pd
    sys.path.insert(0, str(SCRIPT.parent))
    import parse_chats as pc
    x = tmp_path / 'g.xlsx'
    make_xlsx(x, [('数据', '聊天内容', [CHAT])])
    orig = pd.ExcelFile.parse
    calls = {'n': 0}

    def flaky(self, *a, **k):
        calls['n'] += 1
        if calls['n'] > 1:          # pick_sheet 的 nrows=1 试探放行，全量 parse 抛
            raise ValueError('boom')
        return orig(self, *a, **k)

    monkeypatch.setattr(pd.ExcelFile, 'parse', flaky)
    monkeypatch.setattr(sys, 'argv', ['parse_chats.py', str(x), str(tmp_path / 'out')])
    with pytest.raises(SystemExit) as e:
        pc.main()
    assert e.value.code == 1
    assert '读取失败' in capfdbinary.readouterr().out.decode('utf-8')
