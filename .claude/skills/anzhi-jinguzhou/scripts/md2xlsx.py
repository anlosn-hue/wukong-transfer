# -*- coding: utf-8 -*-
"""紧箍咒批次产出 md -> xlsx 转换（02-审计关注点清单 / 03-体检报告）

用法: python md2xlsx.py <批次目录>
读取批次目录下的 02-审计关注点清单.md / 03-体检报告.md，
在同目录生成同名 .xlsx。
排版规则：标题/说明/落款整行合并、落款右对齐、行号/严重度居中、
按内容自适应行高。标题与落款日期均从 md 原文提取，不硬编码。"""
import re
import sys
from math import ceil
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if len(sys.argv) != 2:
    print('用法: python md2xlsx.py <批次目录>')
    sys.exit(2)
BATCH = Path(sys.argv[1])

THIN = Border(*[Side(style='thin', color='BFBFBF')] * 4)
HDR_FILL = PatternFill('solid', fgColor='D9E2F3')
HDR_FONT = Font(bold=True, size=10.5)
BODY_FONT = Font(size=10.5)
WRAP = Alignment(wrap_text=True, vertical='top')
WRAP_C = Alignment(wrap_text=True, vertical='top', horizontal='center')


def vlen(s):
    """可视宽度：CJK 记 2，ASCII 记 1（Excel 列宽单位约等于 1 个 ASCII 字符）"""
    return sum(2 if ord(c) > 127 else 1 for c in str(s))


def fit_height(texts, widths, size=10.5):
    """按各单元格内容与列宽估算该行需要的行高"""
    lines = 1
    for t, w in zip(texts, widths):
        if t is None:
            continue
        eff = max(w - 2, 4)
        n = 0
        for seg in str(t).split('\n'):
            n += max(1, ceil(vlen(seg) / eff))
        lines = max(lines, n)
    return min(lines * (size + 3.5) + 4, 380)


def write_table(ws, rows, row0=1, widths=None, center_cols=()):
    widths = widths or [20] * len(rows[0])
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    for r, cells in enumerate(rows):
        for c, val in enumerate(cells):
            cell = ws.cell(row=row0 + r, column=c + 1, value=val)
            cell.border = THIN
            cell.alignment = WRAP_C if (r == 0 or c in center_cols) else WRAP
            if r == 0:
                cell.font = HDR_FONT
                cell.fill = HDR_FILL
            else:
                cell.font = BODY_FONT
        ws.row_dimensions[row0 + r].height = fit_height(cells, widths)
    return row0 + len(rows)


def add_merged(ws, row, text, ncols, bold=False, size=10.5, align='left', height=True):
    """整行合并写入一段文字"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=bold, size=size)
    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal=align)
    if height:
        total_w = sum(ws.column_dimensions[get_column_letter(c)].width or 8.43
                      for c in range(1, ncols + 1))
        ws.row_dimensions[row].height = fit_height([text], [total_w], size)
    return row + 1


def set_widths(ws, widths):
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w


def parse_table(lines, start):
    rows = []
    i = start
    while i < len(lines) and lines[i].strip().startswith('|'):
        cells = [c.strip() for c in lines[i].strip().strip('|').split('|')]
        if not all(re.fullmatch(r':?-+:?', c) for c in cells):
            rows.append(cells)
        i += 1
    return rows, i


def signoff(lines):
    """md 末尾落款两行（机构名 + 日期）"""
    tail = [l.strip() for l in lines if l.strip()][-2:]
    return tail


def convert_02():
    lines = (BATCH / '02-审计关注点清单.md').read_text(encoding='utf-8').splitlines()
    wb = Workbook()
    ws = wb.active
    ws.title = '审计关注点清单'
    W = [24, 14, 110]
    set_widths(ws, W)
    title = lines[0].lstrip('# ').strip()
    r = add_merged(ws, 1, title, 3, bold=True, size=14)
    for n in [l.lstrip('> ').strip() for l in lines if l.startswith('>')]:
        r = add_merged(ws, r, n, 3)
    r += 1
    tbl_start = next(i for i, l in enumerate(lines) if l.strip().startswith('| 聊天行号'))
    rows, _ = parse_table(lines, tbl_start)
    r = write_table(ws, rows, r, widths=W, center_cols=(1,))
    summary = next(l for l in lines if l.startswith('共 '))
    r = add_merged(ws, r + 1, re.sub(r'\*\*', '', summary), 3)
    s1, s2 = signoff(lines)
    r = add_merged(ws, r + 1, s1, 3, align='right')
    add_merged(ws, r, s2, 3, align='right')
    out = BATCH / '02-审计关注点清单.xlsx'
    wb.save(out)
    return out


def convert_03():
    lines = (BATCH / '03-体检报告.md').read_text(encoding='utf-8').splitlines()
    wb = Workbook()

    def sec(title):
        return lines.index(title) if title in lines else None

    sec7 = sec('## 七、标注结论与质检结论差异核对')
    end6 = sec7 if sec7 is not None else len(lines)

    # ---- Sheet 总览 ----
    ws = wb.active
    ws.title = '总览'
    W = [30, 14, 60]
    set_widths(ws, W)
    NC = 3
    title = lines[0].lstrip('# ').strip()
    r = add_merged(ws, 1, title, NC, bold=True, size=14)
    i = lines.index('## 一、总览')
    for l in lines[1:i]:                      # 标题与首节之间的引言（如复核状态行）
        if l.strip().startswith('>'):
            r = add_merged(ws, r, l.lstrip('> ').strip(), NC)
    j = i + 1
    while not lines[j].startswith('### '):
        s = lines[j].strip()
        if s.startswith('- '):
            r = add_merged(ws, r, s[2:], NC)
        j += 1
    r = add_merged(ws, r + 1, '按维度分布', NC, bold=True)
    k = next(x for x in range(j, len(lines)) if lines[x].strip().startswith('| 维度'))
    rows, k = parse_table(lines, k)
    r = write_table(ws, rows, r, widths=W, center_cols=(1,))
    r = add_merged(ws, r + 1, '按经理分布（有命中者）', NC, bold=True)
    k = next(x for x in range(k, len(lines)) if lines[x].strip().startswith('| 经理'))
    rows, k = parse_table(lines, k)
    # 两列表：经理列 30，命中列合并 B:C 手工写
    for rr, cells in enumerate(rows):
        row_i = r + rr
        c1 = ws.cell(row=row_i, column=1, value=cells[0])
        ws.merge_cells(start_row=row_i, start_column=2, end_row=row_i, end_column=3)
        c2 = ws.cell(row=row_i, column=2, value=cells[1])
        for c in (c1, c2, ws.cell(row=row_i, column=3)):
            c.border = THIN
        if rr == 0:
            c1.font = c2.font = HDR_FONT
            c1.fill = c2.fill = HDR_FILL
            c1.alignment = c2.alignment = WRAP_C
        else:
            c1.font = c2.font = BODY_FONT
            c1.alignment = c2.alignment = WRAP
        ws.row_dimensions[row_i].height = fit_height(cells, [W[0], W[1] + W[2]])
    r = r + len(rows)
    k = lines.index('### 高严重度条目摘要')
    r = add_merged(ws, r + 1, '高严重度条目摘要', NC, bold=True)
    for l in lines[k + 1:]:
        if l.startswith('## '):
            break
        if l.strip():
            r = add_merged(ws, r, l.strip(), NC)
    k = lines.index('## 二、审计关注点')
    r = add_merged(ws, r + 1, '审计关注点', NC, bold=True)
    for l in lines[k + 1:]:
        if l.startswith('## '):
            break
        if l.strip():
            r = add_merged(ws, r, re.sub(r'\*\*', '', l.strip()), NC)

    # ---- Sheet 七维明细 ----
    ws2 = wb.create_sheet('七维明细')
    header = ['维度', '行号', '经理', '严重度', '原话/事项', '规则出处', '改进建议', '标注']
    all_rows = [header]
    sec3 = lines.index('## 三、七维明细')
    sec4 = lines.index('## 四、语音盲区清单（待人工抽听）')
    sec5 = lines.index('## 五、待人工确认项')  # 供语音盲区上界与待人工确认段两处复用
    dim = None
    x = sec3
    while x < sec4:
        l = lines[x]
        if l.startswith('### '):
            dim = l[4:].strip()
        elif l.strip().startswith('|') and '行号' in l:
            rows, x = parse_table(lines, x)
            for row in rows[1:]:
                all_rows.append([dim] + row + [''] * (7 - len(row)))
            continue
        elif dim and l.strip().startswith('无命中'):
            all_rows.append([dim, '-', '-', '-', l.strip(), '', '', ''])
        x += 1
    W2 = [22, 10, 16, 10, 55, 30, 45, 24]
    end = write_table(ws2, all_rows, 1, widths=W2, center_cols=(1, 3))
    ws2.auto_filter.ref = f'A1:H{end - 1}'
    ws2.freeze_panes = 'A2'

    # ---- Sheet 语音盲区 ----
    ws3 = wb.create_sheet('语音盲区待抽听')
    W3 = [10, 18, 45, 60]
    k = next((x for x in range(sec4, sec5) if lines[x].strip().startswith('| 行号')), None)
    if k is None:      # 无待抽听表：只写一行说明，不留空 sheet
        set_widths(ws3, W3)
        note = next((l.strip() for l in lines[sec4 + 1:sec5]
                     if l.strip() and not l.startswith('### ')), '无待抽听项')
        add_merged(ws3, 1, note, 4)
    else:
        rows, k = parse_table(lines, k)
        end = write_table(ws3, rows, 1, widths=W3, center_cols=(0,))
        for l in lines[k:]:
            if l.startswith('## '):
                break
            if l.strip().startswith('注：'):
                add_merged(ws3, end + 1, l.strip(), 4)
        ws3.freeze_panes = 'A2'

    # ---- Sheet 待人工确认 ----
    ws4 = wb.create_sheet('待人工确认')
    sec6 = lines.index('## 六、沉淀建议')
    rows = [['#', '事项']]
    for l in lines[sec5 + 1:sec6]:
        if l.startswith('### '):      # 抽检建议等三级小节及其后内容不进本 sheet
            break
        m = re.match(r'(\d+)\.\s+(.*)', l.strip())
        if m:
            rows.append([m.group(1), re.sub(r'\*\*', '', m.group(2))])
    write_table(ws4, rows, 1, widths=[6, 120], center_cols=(0,))

    # ---- Sheet 沉淀建议 ----
    ws5 = wb.create_sheet('沉淀建议')
    rows = [['#', '建议']]
    for l in lines[sec6 + 1:end6]:
        m = re.match(r'(\d+)\.\s+(.*)', l.strip())
        if m:
            rows.append([m.group(1), re.sub(r'\*\*', '', m.group(2))])
    end = write_table(ws5, rows, 1, widths=[6, 120], center_cols=(0,))
    s1, s2 = signoff(lines)
    end = add_merged(ws5, end + 1, s1, 2, align='right')
    add_merged(ws5, end, s2, 2, align='right')

    # ---- Sheet 差异核对（03 无第七节的老批次兼容跳过）----
    if sec7 is not None:
        ws6 = wb.create_sheet('差异核对')
        W6 = [10, 30, 30, 22, 45]
        try:
            k = next(x for x in range(sec7, len(lines))
                     if lines[x].strip().startswith('| 行号'))
            rows, _ = parse_table(lines, k)
            write_table(ws6, rows, 1, widths=W6, center_cols=(0,))
            ws6.freeze_panes = 'A2'
        except StopIteration:      # 零差异批次：只有说明行
            set_widths(ws6, W6)
            nonblank = [i for i, l in enumerate(lines) if l.strip()]
            sig_start = nonblank[-2] if len(nonblank) >= 2 else len(lines)
            note = next((l.strip() for l in lines[sec7 + 1:sig_start]
                         if l.strip() and not l.startswith('## ')), '本批无实质差异')
            add_merged(ws6, 1, note, 5)

    out = BATCH / '03-体检报告.xlsx'
    wb.save(out)
    return out


if __name__ == '__main__':
    p1 = convert_02()
    p2 = convert_03()
    sys.stdout.buffer.write(f'OK\n{p1}\n{p2}\n'.encode('utf-8'))
