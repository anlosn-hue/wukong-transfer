# -*- coding: utf-8 -*-
"""紧箍咒 Step 3 产出合成：04-标注.jsonl + 01-原文映射.jsonl -> 04-合规预警标注表.xlsx

三道闸门（任一不过：报错列明细、exit 1、不生成 xlsx）：
1. 完整性：标注行号集合 ≡ 原文映射行号集合（缺行=漏判，多行=幻觉行号），且两侧行号均无重复
2. 格式：problem ∈ {是,否}；四数组字段须为数组；判是行四数组等长非空、无空字符串；
   判否行 labels/originals/confidence 为空、reasons 恰一条非空
3. 枚举：label ∈ 12 正式名（拦样例异名），confidence ∈ 1—5 级
用法: python label2xlsx.py <批次目录>"""
import json
import sys
from math import ceil
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

LABELS = {
    '保本保息或收益承诺', '宣传预期收益率或未来收益', '风险提示不到位或弱化风险',
    '未宣告业绩比较基准、过往业绩不代表未来收益', '未说明代销性质',
    '混淆产品性质或不当类比', '极端或夸大性表述', '饥饿营销', '不当利益诱导营销',
    '贬低同业产品或机构', '利用监管、政府或公共机构公信力营销', '骚扰客户',
}
LEVELS = {'1级', '2级', '3级', '4级', '5级'}
HEADERS = ['聊天内容原文', '是否存在问题', '违规预警标签',
           '问题原文', '判断理由', '风险判断可信度等级']
WIDTHS = [80, 10, 28, 45, 55, 16]


def say(msg):
    sys.stdout.buffer.write((msg + '\n').encode('utf-8'))


def load_jsonl(path):
    return [json.loads(l) for l in path.read_text(encoding='utf-8').splitlines()
            if l.strip()]


def check(ann, mapping):
    errs = []
    mrows = [m.get('row') for m in mapping]
    mdup = sorted({x for x in mrows if mrows.count(x) > 1})
    if mdup:
        errs.append(f'原文映射行号重复：{mdup}')
    rows = [r.get('row') for r in ann]
    dup = sorted({x for x in rows if rows.count(x) > 1})
    if dup:
        errs.append(f'行号重复：{dup}')
    aset, mset = set(rows), set(mrows)
    if mset - aset:
        errs.append(f'缺行（漏判）：{sorted(mset - aset)}')
    if aset - mset:
        errs.append(f'多行（幻觉行号）：{sorted(aset - mset)}')
    for r in ann:
        w = f'行{r.get("row")}'
        if r.get('problem') not in ('是', '否'):
            errs.append(f'{w}：problem 非法值「{r.get("problem")}」')
            continue
        L, O, R, C = (r.get(k) or [] for k in
                      ('labels', 'originals', 'reasons', 'confidence'))
        if not all(isinstance(x, list) for x in (L, O, R, C)):
            errs.append(f'{w}：labels/originals/reasons/confidence 须为数组')
            continue
        if r['problem'] == '是':
            # 四数组的位置对应关系（labels[i]↔originals[i]↔reasons[i]↔confidence[i]）
            # 是上游 LLM 标注的语义约定，机械闸门只验数量（等长非空）不验对应，信任边界在此。
            if not (len(L) == len(O) == len(R) == len(C) >= 1):
                errs.append(f'{w}：判是行四数组须等长且非空')
                continue
            if any(not str(x).strip() for x in L + O + R + C):
                errs.append(f'{w}：判是行含空字符串')
            bad = [x for x in L if x not in LABELS]
            if bad:
                errs.append(f'{w}：非法标签（非12正式名）：{bad}')
            badc = [x for x in C if x not in LEVELS]
            if badc:
                errs.append(f'{w}：可信度非法值：{badc}')
        else:
            if L or O or C:
                errs.append(f'{w}：判否行 labels/originals/confidence 必须为空')
            if len(R) != 1 or not str(R[0]).strip():
                errs.append(f'{w}：判否行 reasons 须恰一条非空')
    return errs


def numbered(labels, items):
    return '；'.join(f'{i}.{l}：{v}' for i, (l, v) in
                     enumerate(zip(labels, items), 1))


def vlen(s):
    return sum(2 if ord(c) > 127 else 1 for c in str(s))


def fit_height(texts, widths, size=10.5):
    lines = 1
    for t, w in zip(texts, widths):
        if not t:
            continue
        eff = max(w - 2, 4)
        n = sum(max(1, ceil(vlen(seg) / eff)) for seg in str(t).split('\n'))
        lines = max(lines, n)
    return min(lines * (size + 3.5) + 4, 380)


def main():
    if len(sys.argv) != 2:
        say('用法: python label2xlsx.py <批次目录>')
        sys.exit(2)
    batch = Path(sys.argv[1])
    ann_path = batch / '04-标注.jsonl'
    mapping_path = batch / '01-原文映射.jsonl'
    try:
        ann = load_jsonl(ann_path)
    except (OSError, json.JSONDecodeError) as e:
        say(f'格式异常：{ann_path.name} 读不到或非法 JSON（{e}），停止。请报告用户。')
        sys.exit(1)
    try:
        mapping = load_jsonl(mapping_path)
    except (OSError, json.JSONDecodeError) as e:
        say(f'格式异常：{mapping_path.name} 读不到或非法 JSON（{e}），停止。请报告用户。')
        sys.exit(1)
    errs = check(ann, mapping)
    if errs:
        say(f'闸门不过（{len(errs)} 项），不生成 xlsx：')
        for e in errs:
            say('  - ' + e)
        sys.exit(1)
    raw = {m['row']: m['raw'] for m in mapping}
    wb = Workbook()
    ws = wb.active
    ws.title = '标注表'
    thin = Border(*[Side(style='thin', color='BFBFBF')] * 4)
    for c, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
        ws.column_dimensions[get_column_letter(c)].width = w
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True, size=10.5)
        cell.fill = PatternFill('solid', fgColor='D9E2F3')
        cell.alignment = Alignment(wrap_text=True, vertical='center',
                                   horizontal='center')
        cell.border = thin
    for i, r in enumerate(sorted(ann, key=lambda x: x['row']), 2):
        if r['problem'] == '是':
            vals = [raw[r['row']], '是', '；'.join(r['labels']),
                    numbered(r['labels'], r['originals']),
                    numbered(r['labels'], r['reasons']),
                    '；'.join(r['confidence'])]
        else:
            vals = [raw[r['row']], '否', None, None, r['reasons'][0], None]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(i, c, v)
            cell.font = Font(size=10.5)
            cell.border = thin
            cell.alignment = Alignment(
                wrap_text=True, vertical='top',
                horizontal='center' if c in (2, 6) else 'left')
        ws.row_dimensions[i].height = fit_height(vals, WIDTHS)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:F{len(ann) + 1}'
    out = batch / '04-合规预警标注表.xlsx'
    wb.save(out)
    say(f'OK 共{len(ann)}行（判是{sum(1 for r in ann if r["problem"] == "是")}）\n{out}')


if __name__ == '__main__':
    main()
